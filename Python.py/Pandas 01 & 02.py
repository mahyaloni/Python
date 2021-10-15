import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
df=pd.DataFrame({
    'year':[1383,1384,1385,1386,1387,1388,1389,1390,1391,1392,1393,1394,1395,1396,1397,1398],
    'GDP':[4945369,5199769,5476337,5844885,5840480,5840800,6175274,6364368,5873423,5854329,6042535,5946680,6691109,6940834,6564709,6137336],
    'capital':[14663907,15294782,15868474,16584954,17435801,18273737,19127886,20017053,20541134,20927962,21388759,21658999,21863000,22063661,22101762,22071910]},
     index=[0,1,2,3,4,5,6,7,8,9,10,11,12,13,14,15])

df['percent']=df['capital']/df['GDP']*100
print(df['percent'])

plt.figure(figsize=(15,8))

plt.subplot(2,3,1)
plt.plot(df['percent'])
plt.grid()
plt.show()

import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import load_workbook
book=load_workbook('GDP Per Capita.xlsx')
sheet=book.active
x=np.zeros((31,5))
for i in range(31):
    for j in range(5):
        x[i,j]=sheet.cell(row=i+2,column=j+1).value
print(x)
plt.plot(x[:,0],x[:,1],label='Iran')
plt.plot(x[:,0],x[:,2],label='Saudi Arabia')
plt.plot(x[:,0],x[:,3],label='Turky')
plt.plot(x[:,0],x[:,4],label='US')
plt.xlabel('year')
plt.ylabel('GDP')
plt.legend()
plt.grid()
plt.show()

# output
0     296.517954
1     294.143490
2     289.764381
3     283.751588
4     298.533699
5     312.863597
6     309.749592
7     314.517529
8     349.730200
9     357.478406
10    353.969965
11    364.220019
12    326.747031
13    317.881987
14    336.675426
15    359.633398





